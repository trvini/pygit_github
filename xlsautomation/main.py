#test python script for xls sheet automation
import openpyxl
from openpyxl.styles import Font,PatternFill

Ratings_File = openpyxl.load_workbook("NoteRating.xlsx")
Rating_List = Ratings_File["Sheet1"]

#create new worksheet
Newsheet = Ratings_File.create_sheet("Mysheet", 0) # insert at the end (default)
Newsheet.title = "Dashboard"

top_left_cellA = Newsheet['A1']
top_left_cellA.font= Font(b=True)
top_left_cellA.fill = PatternFill("solid", fgColor="00FFCC99")
top_left_cellA.value = "Name"

top_left_cellB = Newsheet['B1']
top_left_cellB.font= Font(b=True)
top_left_cellB.fill = PatternFill("solid", fgColor="00FFCC99")
top_left_cellB.value = "#of tests"

# sheet.maxrow will give the info of how many rows are filled uo in the sheet
# to run a for loop we need to cretae a list of numbers inorder to loop thru each item in that list
# to get that list we need to use range funtion.
print(range(Rating_List.max_row))

# range starts from value 0, our sheet has no 0, start reading the rows only from row 2 since row 1 is title
# create a dictionary with key-value pair -  childname as key and number of tests as value
tests_per_child = {}

for each_row in range(2, Rating_List.max_row+1):
    Child_name = Rating_List.cell(each_row,1).value


    if Child_name in tests_per_child:
        tests_per_child[Child_name] += 1
    else:
        tests_per_child[Child_name] = 1
        Newsheet.cell(each_row, 1).value = Child_name

for i in range(1, len(tests_per_child)+1):
    Child_name = Newsheet.cell(i+1, 1).value
    Newsheet.cell(i+1, 2).value = tests_per_child.get(Child_name)

print(len(tests_per_child))
print(tests_per_child)



Ratings_File.save("NoteRating.xlsx")